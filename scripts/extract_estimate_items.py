from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:  # pragma: no cover - optional dependency for old .xls files.
    xlrd = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_ROOT = PROJECT_ROOT / "data" / "raw_estimates"
PROCESSED_ROOT = PROJECT_ROOT / "data" / "processed"
CACHE_PATH = PROCESSED_ROOT / "estimate_extract_cache.json"
CACHE_VERSION = 2

LABEL_RECIPIENT = "\uc218\uc2e0"
LABEL_ATTENTION = "\ucc38\uc870"
LABEL_QUOTE_TITLE = "\uacac\uc801\uba85"
LABEL_QUOTE_DATE = "\uc77c\uc790"


ITEM_FIELDS = [
    "source_path",
    "source_file",
    "source_folder",
    "customer_hint",
    "quote_date",
    "recipient",
    "attention",
    "quote_title",
    "sheet_name",
    "item_no",
    "item_category",
    "part_name",
    "spec",
    "quantity",
    "unit",
    "quoted_unit_price",
    "amount",
    "normalized_part_key",
]

FILE_FIELDS = [
    "source_path",
    "source_file",
    "source_folder",
    "customer_hint",
    "file_extension",
    "file_modified_at",
    "status",
    "item_count",
    "note",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).replace("\xa0", " ").replace("\n", " ").strip()


def parse_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    return clean(value).replace(",", "")


def normalize_key(*parts: str) -> str:
    text = " ".join(p for p in parts if p).lower()
    text = re.sub(r"[^0-9a-zA-Z\uac00-\ud7a3]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def date_from_filename(name: str) -> str:
    match = re.search(r"(20\d{2})[-_. ]?(\d{2})[-_. ]?(\d{2})", name)
    if not match:
        return ""
    year, month, day = match.groups()
    try:
        return datetime(int(year), int(month), int(day)).date().isoformat()
    except ValueError:
        return ""


def file_signature(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
        "suffix": path.suffix.lower(),
    }


def make_file_record(path: Path, status: str = "ok", note: str = "", item_count: int = 0) -> dict[str, str]:
    rel_path = path.relative_to(RAW_ROOT)
    folder_parts = rel_path.parts[:-1]
    return {
        "source_path": str(rel_path),
        "source_file": path.name,
        "source_folder": "\\".join(folder_parts),
        "customer_hint": folder_parts[-1] if folder_parts else "",
        "file_extension": path.suffix.lower(),
        "file_modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        "status": status,
        "item_count": str(item_count),
        "note": note,
    }


def load_cache() -> dict[str, Any]:
    if CACHE_PATH.exists():
        with CACHE_PATH.open("r", encoding="utf-8") as file:
            cache = json.load(file)
        if cache.get("version") == CACHE_VERSION:
            return cache
    return {"version": CACHE_VERSION, "files": {}}


def save_cache(cache: dict[str, Any]) -> None:
    tmp_path = CACHE_PATH.with_suffix(".tmp")
    with tmp_path.open("w", encoding="utf-8") as file:
        json.dump(cache, file, ensure_ascii=False)
    tmp_path.replace(CACHE_PATH)


def find_label_value(ws, label: str, max_row: int = 20, max_col: int = 10) -> str:
    normalized = label.replace(" ", "")
    for row in range(1, min(ws.max_row, max_row) + 1):
        for col in range(1, min(ws.max_column, max_col) + 1):
            value = clean(ws.cell(row, col).value).replace(" ", "")
            if value == normalized:
                for offset in range(1, 3):
                    candidate = clean(ws.cell(row, col + offset).value)
                    if candidate:
                        return candidate
    return ""


def find_quote_date(ws) -> str:
    for row in range(1, min(ws.max_row, 20) + 1):
        for col in range(1, min(ws.max_column, 10) + 1):
            value = clean(ws.cell(row, col).value).replace(" ", "")
            if value in {LABEL_QUOTE_DATE, f"{LABEL_QUOTE_DATE}:"}:
                for offset in range(1, 5):
                    text = clean(ws.cell(row, col + offset).value)
                    if text:
                        return text
    return ""


def find_header(ws) -> tuple[int, int] | None:
    for row in range(1, min(ws.max_row, 60) + 1):
        for col in range(1, min(ws.max_column, 20) + 1):
            if clean(ws.cell(row, col).value) == "No.":
                return row, col
    return None


def extract_xlsx(path: Path) -> tuple[list[dict[str, str]], dict[str, str]]:
    rows: list[dict[str, str]] = []
    rel_path = path.relative_to(RAW_ROOT)
    folder_parts = rel_path.parts[:-1]
    customer_hint = folder_parts[-1] if folder_parts else ""
    file_record = make_file_record(path)

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return rows, make_file_record(path, status="error", note=f"{type(exc).__name__}: {exc}")

    try:
        for ws in workbook.worksheets:
            header = find_header(ws)
            if not header:
                continue

            header_row, no_col = header
            recipient = find_label_value(ws, LABEL_RECIPIENT)
            attention = find_label_value(ws, LABEL_ATTENTION)
            quote_title = find_label_value(ws, LABEL_QUOTE_TITLE)
            quote_date = find_quote_date(ws) or date_from_filename(path.name)

            for row_num in range(header_row + 1, min(ws.max_row, 300) + 1):
                item_no = clean(ws.cell(row_num, no_col).value)
                item_category = clean(ws.cell(row_num, no_col + 1).value)
                spec = clean(ws.cell(row_num, no_col + 2).value)
                quantity = parse_number(ws.cell(row_num, no_col + 3).value)
                unit = clean(ws.cell(row_num, no_col + 4).value)
                unit_price = parse_number(ws.cell(row_num, no_col + 5).value)
                amount = parse_number(ws.cell(row_num, no_col + 6).value)

                if not item_no.isdigit() or (not item_category and not spec):
                    continue

                rows.append(
                    {
                        "source_path": str(rel_path),
                        "source_file": path.name,
                        "source_folder": "\\".join(folder_parts),
                        "customer_hint": customer_hint,
                        "quote_date": quote_date,
                        "recipient": recipient,
                        "attention": attention,
                        "quote_title": quote_title,
                        "sheet_name": ws.title,
                        "item_no": item_no,
                        "item_category": item_category,
                        "part_name": spec,
                        "spec": spec,
                        "quantity": quantity,
                        "unit": unit,
                        "quoted_unit_price": unit_price,
                        "amount": amount,
                        "normalized_part_key": normalize_key(item_category, spec),
                    }
                )
    finally:
        workbook.close()

    status = "ok" if rows else "no_items"
    file_record["status"] = status
    file_record["item_count"] = str(len(rows))
    return rows, file_record


def xls_value(workbook, sheet, row: int, col: int) -> Any:
    if row < 0 or col < 0 or row >= sheet.nrows or col >= sheet.ncols:
        return None
    cell = sheet.cell(row, col)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return datetime(*xlrd.xldate_as_tuple(cell.value, workbook.datemode))
        except Exception:
            return cell.value
    return cell.value


def find_header_xls(sheet) -> tuple[int, int] | None:
    for row in range(0, min(sheet.nrows, 60)):
        for col in range(0, min(sheet.ncols, 20)):
            if clean(sheet.cell_value(row, col)) == "No.":
                return row, col
    return None


def find_label_value_xls(workbook, sheet, label: str, max_row: int = 20, max_col: int = 10) -> str:
    normalized = label.replace(" ", "")
    for row in range(0, min(sheet.nrows, max_row)):
        for col in range(0, min(sheet.ncols, max_col)):
            value = clean(xls_value(workbook, sheet, row, col)).replace(" ", "")
            if value == normalized:
                for offset in range(1, 3):
                    candidate = clean(xls_value(workbook, sheet, row, col + offset))
                    if candidate:
                        return candidate
    return ""


def find_quote_date_xls(workbook, sheet) -> str:
    for row in range(0, min(sheet.nrows, 20)):
        for col in range(0, min(sheet.ncols, 10)):
            value = clean(xls_value(workbook, sheet, row, col)).replace(" ", "")
            if value in {LABEL_QUOTE_DATE, f"{LABEL_QUOTE_DATE}:"}:
                for offset in range(1, 5):
                    text = clean(xls_value(workbook, sheet, row, col + offset))
                    if text:
                        return text
    return ""


def extract_xls(path: Path) -> tuple[list[dict[str, str]], dict[str, str]]:
    rows: list[dict[str, str]] = []
    if xlrd is None:
        return rows, make_file_record(path, status="unsupported_pending", note="Install xlrd to extract .xls files.")

    rel_path = path.relative_to(RAW_ROOT)
    folder_parts = rel_path.parts[:-1]
    customer_hint = folder_parts[-1] if folder_parts else ""
    file_record = make_file_record(path)

    try:
        workbook = xlrd.open_workbook(str(path))
    except Exception as exc:
        return rows, make_file_record(path, status="error", note=f"{type(exc).__name__}: {exc}")

    for sheet in workbook.sheets():
        header = find_header_xls(sheet)
        if not header:
            continue

        header_row, no_col = header
        recipient = find_label_value_xls(workbook, sheet, LABEL_RECIPIENT)
        attention = find_label_value_xls(workbook, sheet, LABEL_ATTENTION)
        quote_title = find_label_value_xls(workbook, sheet, LABEL_QUOTE_TITLE)
        quote_date = find_quote_date_xls(workbook, sheet) or date_from_filename(path.name)

        for row_num in range(header_row + 1, min(sheet.nrows, 300)):
            item_no = clean(xls_value(workbook, sheet, row_num, no_col))
            if item_no.endswith(".0"):
                item_no = item_no[:-2]
            item_category = clean(xls_value(workbook, sheet, row_num, no_col + 1))
            spec = clean(xls_value(workbook, sheet, row_num, no_col + 2))
            quantity = parse_number(xls_value(workbook, sheet, row_num, no_col + 3))
            unit = clean(xls_value(workbook, sheet, row_num, no_col + 4))
            unit_price = parse_number(xls_value(workbook, sheet, row_num, no_col + 5))
            amount = parse_number(xls_value(workbook, sheet, row_num, no_col + 6))

            if not item_no.isdigit() or (not item_category and not spec):
                continue

            rows.append(
                {
                    "source_path": str(rel_path),
                    "source_file": path.name,
                    "source_folder": "\\".join(folder_parts),
                    "customer_hint": customer_hint,
                    "quote_date": quote_date,
                    "recipient": recipient,
                    "attention": attention,
                    "quote_title": quote_title,
                    "sheet_name": sheet.name,
                    "item_no": item_no,
                    "item_category": item_category,
                    "part_name": spec,
                    "spec": spec,
                    "quantity": quantity,
                    "unit": unit,
                    "quoted_unit_price": unit_price,
                    "amount": amount,
                    "normalized_part_key": normalize_key(item_category, spec),
                }
            )

    status = "ok" if rows else "no_items"
    file_record["status"] = status
    file_record["item_count"] = str(len(rows))
    return rows, file_record


def get_cached(cache: dict[str, Any], path: Path) -> tuple[list[dict[str, str]], dict[str, str]] | None:
    rel_path = str(path.relative_to(RAW_ROOT))
    entry = cache.get("files", {}).get(rel_path)
    if not entry:
        return None
    if entry.get("signature") != file_signature(path):
        return None
    return entry.get("items", []), entry.get("record", make_file_record(path))


def put_cached(
    cache: dict[str, Any],
    path: Path,
    items: list[dict[str, str]],
    record: dict[str, str],
) -> None:
    rel_path = str(path.relative_to(RAW_ROOT))
    cache.setdefault("files", {})[rel_path] = {
        "signature": file_signature(path),
        "items": items,
        "record": record,
    }


def extract_path(path: Path, cache: dict[str, Any]) -> tuple[list[dict[str, str]], dict[str, str], bool]:
    cached = get_cached(cache, path)
    if cached:
        items, record = cached
        return items, record, True

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        items, record = extract_xlsx(path)
    elif suffix == ".xls":
        items, record = extract_xls(path)
    elif suffix in {".xlsm", ".csv"}:
        items = []
        record = make_file_record(
            path,
            status="unsupported_pending",
            note="This file type is pending extraction support.",
        )
    else:
        items = []
        record = make_file_record(path, status="ignored", note="Unsupported file extension.")

    put_cached(cache, path, items, record)
    return items, record, False


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    PROCESSED_ROOT.mkdir(parents=True, exist_ok=True)

    cache = load_cache()
    all_items: list[dict[str, str]] = []
    file_records: list[dict[str, str]] = []
    cache_hits = 0
    processed = 0

    current_paths = set()
    for path in sorted(RAW_ROOT.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in {".xlsx", ".xls", ".xlsm", ".csv"}:
            continue

        rel_path = str(path.relative_to(RAW_ROOT))
        current_paths.add(rel_path)
        items, record, was_cached = extract_path(path, cache)
        all_items.extend(items)
        file_records.append(record)
        if was_cached:
            cache_hits += 1
        else:
            processed += 1

    for rel_path in list(cache.get("files", {}).keys()):
        if rel_path not in current_paths:
            del cache["files"][rel_path]

    write_csv(PROCESSED_ROOT / "estimate_items.csv", ITEM_FIELDS, all_items)
    write_csv(PROCESSED_ROOT / "estimate_files.csv", FILE_FIELDS, file_records)
    save_cache(cache)

    print(f"files scanned: {len(file_records)}")
    print(f"cache hits: {cache_hits}")
    print(f"files processed: {processed}")
    print(f"items extracted: {len(all_items)}")
    print(f"output: {PROCESSED_ROOT}")


if __name__ == "__main__":
    main()
