from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PROCESSED_ROOT = PROJECT_ROOT / "data" / "processed"
ITEMS_CSV = PROCESSED_ROOT / "estimate_items.csv"
OUTPUT_XLSX = PROCESSED_ROOT / "estimate_search.xlsx"


COLUMNS = [
    ("quote_date", "견적일"),
    ("customer_hint", "폴더/업체힌트"),
    ("recipient", "수신"),
    ("attention", "참조"),
    ("item_category", "품목"),
    ("spec", "부품명/규격"),
    ("quantity", "수량"),
    ("unit", "단위"),
    ("quoted_unit_price", "견적단가"),
    ("amount", "금액"),
    ("source_file", "원본파일"),
    ("source_path", "원본경로"),
    ("normalized_part_key", "검색키"),
]


def to_number(value: str):
    text = (value or "").replace(",", "").strip()
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return value
    if number.is_integer():
        return int(number)
    return number


def main() -> None:
    if not ITEMS_CSV.exists():
        raise SystemExit(f"Missing data file: {ITEMS_CSV}")

    with ITEMS_CSV.open("r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.DictReader(file))

    workbook = Workbook()
    ws = workbook.active
    ws.title = "견적품목검색"

    ws.append([title for _, title in COLUMNS])

    for row in rows:
        values = []
        for key, _ in COLUMNS:
            value = row.get(key, "")
            if key in {"quantity", "quoted_unit_price", "amount"}:
                value = to_number(value)
            values.append(value)
        ws.append(values)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 12,
        "B": 16,
        "C": 16,
        "D": 16,
        "E": 10,
        "F": 56,
        "G": 8,
        "H": 8,
        "I": 12,
        "J": 12,
        "K": 34,
        "L": 54,
        "M": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col in ("G", "I", "J"):
        for cell in ws[col][1:]:
            cell.number_format = "#,##0"

    summary = workbook.create_sheet("요약")
    summary.append(["항목", "값"])
    summary.append(["품목 행 수", len(rows)])
    summary.append(["원본 CSV", str(ITEMS_CSV)])
    summary.append(["사용법", "견적품목검색 시트에서 Ctrl+F 또는 필터로 부품명/업체/단가를 검색"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 90

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    print(f"created: {OUTPUT_XLSX}")
    print(f"rows: {len(rows)}")


if __name__ == "__main__":
    main()

