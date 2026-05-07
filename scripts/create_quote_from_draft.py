from __future__ import annotations

import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import win32com.client
except ImportError:  # pragma: no cover - optional Windows Excel automation.
    win32com = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_ROOT = PROJECT_ROOT / "data" / "raw_estimates"
EXPORT_ROOT = PROJECT_ROOT / "data" / "exports"
RAW_QUOTE_ROOT = RAW_ROOT / "\uacac\uc801\uc11c"

LABEL_RECIPIENT = "\uc218\uc2e0"
LABEL_ATTENTION = "\ucc38\uc870"
LABEL_QUOTE_TITLE = "\uacac\uc801\uba85"
LABEL_QUOTE_DATE = "\uc77c\uc790"
LABEL_TOTAL = "\uc77c\uae08"
VAT_SPLIT_LABEL = "\ubd80\uac00\uc138"


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def to_number(value: Any) -> int:
    try:
        return int(float(str(value or "0").replace(",", "")))
    except ValueError:
        return 0


def safe_filename(value: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', "_", value)
    text = re.sub(r"\s+", "_", text).strip("._")
    return text or "quote"


def find_header(ws) -> tuple[int, int] | None:
    for row in range(1, min(ws.max_row, 80) + 1):
        for col in range(1, min(ws.max_column, 25) + 1):
            if clean(ws.cell(row, col).value) == "No.":
                return row, col
    return None


def workbook_has_vat_split(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                header = find_header(ws)
                if header and has_vat_split_columns(ws, *header):
                    return True
        finally:
            wb.close()
    except Exception:
        return False
    return False


def find_template(prefer_vat_split: bool = True) -> Path:
    explicit = PROJECT_ROOT / "data" / "templates" / "quote_template.xlsx"
    if explicit.exists():
        return explicit

    paths = [path for path in sorted(RAW_ROOT.rglob("*.xlsx")) if not path.name.startswith("~$")]
    if prefer_vat_split:
        for path in paths:
            if workbook_has_vat_split(path):
                return path

    for path in paths:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                if any(find_header(ws) for ws in wb.worksheets):
                    return path
            finally:
                wb.close()
        except Exception:
            continue

    raise SystemExit("No usable .xlsx quote template found.")


def find_label_cell(ws, label: str, max_row: int = 20, max_col: int = 12):
    normalized = label.replace(" ", "")
    for row in range(1, min(ws.max_row, max_row) + 1):
        for col in range(1, min(ws.max_column, max_col) + 1):
            value = clean(ws.cell(row, col).value).replace(" ", "")
            if value == normalized:
                return row, col
    return None


def set_near_label(ws, label: str, value: Any, offset: int = 1) -> None:
    cell = find_label_cell(ws, label)
    if cell:
        row, col = cell
        ws.cell(row, col + offset).value = value


def clear_old_items(ws, header_row: int, no_col: int) -> None:
    for row in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
        for col in range(no_col, no_col + 10):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def set_value(ws, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def has_vat_split_columns(ws, header_row: int, no_col: int) -> bool:
    header_text = " ".join(clean(ws.cell(header_row, col).value) for col in range(no_col, no_col + 10))
    return VAT_SPLIT_LABEL in header_text


def unit_price_from_item(item: dict[str, Any]) -> int:
    if item.get("tax_excluded_price"):
        return to_number(item.get("tax_excluded_price"))
    if item.get("price"):
        return to_number(item.get("price"))
    if item.get("quoted_unit_price"):
        return to_number(item.get("quoted_unit_price"))
    if item.get("tax_included_price"):
        return included_to_excluded(to_number(item.get("tax_included_price")), mode="nearest")
    return 0


def included_to_excluded(included_price: int, mode: str = "nearest", unit: int = 1000) -> int:
    if included_price <= 0:
        return 0
    exact = included_price / 1.1
    lower = int(exact // unit) * unit
    upper = lower if lower == exact else lower + unit
    if mode == "floor":
        return lower
    if mode == "ceil":
        return upper
    if abs((upper * 1.1) - included_price) < abs(included_price - (lower * 1.1)):
        return upper
    return lower


def write_items(ws, header_row: int, no_col: int, items: list[dict[str, Any]]) -> int:
    total = 0
    vat_split = has_vat_split_columns(ws, header_row, no_col)
    for index, item in enumerate(items, start=1):
        row = header_row + 1 + index
        qty = to_number(item.get("qty") or item.get("quantity") or 1)
        price = unit_price_from_item(item)
        vat = int(round(price * 0.1))
        amount = qty * (price + vat) if vat_split else qty * price
        total += amount

        set_value(ws, row, no_col, index)
        set_value(ws, row, no_col + 1, item.get("category") or item.get("item_category") or "")
        set_value(ws, row, no_col + 2, item.get("spec") or "")
        set_value(ws, row, no_col + 3, qty)
        set_value(ws, row, no_col + 4, item.get("unit") or "EA")
        set_value(ws, row, no_col + 5, price)
        if vat_split:
            set_value(ws, row, no_col + 6, f"={ws.cell(row, no_col + 5).coordinate}*0.1")
            set_value(
                ws,
                row,
                no_col + 7,
                f"=({ws.cell(row, no_col + 5).coordinate}+{ws.cell(row, no_col + 6).coordinate})*{ws.cell(row, no_col + 3).coordinate}",
            )
            if item.get("tax_included_price"):
                set_value(ws, row, no_col + 8, to_number(item.get("tax_included_price")))
        else:
            set_value(ws, row, no_col + 6, amount)
    return total


def cell_text_com(ws, row: int, col: int) -> str:
    value = ws.Cells(row, col).Value
    return clean(value)


def find_header_com(ws) -> tuple[int, int] | None:
    for row in range(1, 81):
        for col in range(1, 26):
            if cell_text_com(ws, row, col) == "No.":
                return row, col
    return None


def has_vat_split_columns_com(ws, header_row: int, no_col: int) -> bool:
    header_text = " ".join(cell_text_com(ws, header_row, col) for col in range(no_col, no_col + 10))
    return VAT_SPLIT_LABEL in header_text


def set_value_com(ws, row: int, col: int, value: Any) -> None:
    try:
        ws.Cells(row, col).Value = value
    except Exception:
        pass


def set_formula_com(ws, row: int, col: int, formula: str) -> None:
    try:
        ws.Cells(row, col).Formula = formula
    except Exception:
        pass


def clear_old_items_com(ws, header_row: int, no_col: int) -> None:
    for row in range(header_row + 1, header_row + 81):
        for col in range(no_col, no_col + 10):
            try:
                ws.Cells(row, col).ClearContents()
            except Exception:
                pass


def find_label_cell_com(ws, label: str) -> tuple[int, int] | None:
    normalized = label.replace(" ", "")
    for row in range(1, 21):
        for col in range(1, 13):
            if cell_text_com(ws, row, col).replace(" ", "") == normalized:
                return row, col
    return None


def set_near_label_com(ws, label: str, value: Any, offset: int = 1) -> None:
    cell = find_label_cell_com(ws, label)
    if cell:
        row, col = cell
        set_value_com(ws, row, col + offset, value)


def write_items_com(ws, header_row: int, no_col: int, items: list[dict[str, Any]]) -> int:
    total = 0
    vat_split = has_vat_split_columns_com(ws, header_row, no_col)
    for index, item in enumerate(items, start=1):
        row = header_row + 1 + index
        qty = to_number(item.get("qty") or item.get("quantity") or 1)
        price = unit_price_from_item(item)
        vat = int(round(price * 0.1))
        amount = qty * (price + vat) if vat_split else qty * price
        total += amount

        set_value_com(ws, row, no_col, index)
        set_value_com(ws, row, no_col + 1, item.get("category") or item.get("item_category") or "")
        set_value_com(ws, row, no_col + 2, item.get("spec") or "")
        set_value_com(ws, row, no_col + 3, qty)
        set_value_com(ws, row, no_col + 4, item.get("unit") or "EA")
        set_value_com(ws, row, no_col + 5, price)
        if vat_split:
            unit_addr = f"{get_column_letter(no_col + 5)}{row}"
            vat_addr = f"{get_column_letter(no_col + 6)}{row}"
            qty_addr = f"{get_column_letter(no_col + 3)}{row}"
            set_formula_com(ws, row, no_col + 6, f"={unit_addr}*0.1")
            set_formula_com(
                ws,
                row,
                no_col + 7,
                f"=({unit_addr}+{vat_addr})*{qty_addr}",
            )
            if item.get("tax_included_price"):
                set_value_com(ws, row, no_col + 8, to_number(item.get("tax_included_price")))
        else:
            set_value_com(ws, row, no_col + 6, amount)
    return total


def create_quote_with_excel_com(output: Path, payload: dict[str, Any], items: list[dict[str, Any]]) -> bool:
    if win32com is None:
        return False

    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(output.resolve()))
        ws = workbook.Worksheets(1)
        header = find_header_com(ws)
        if not header:
            return False

        header_row, no_col = header
        clear_old_items_com(ws, header_row, no_col)
        total = write_items_com(ws, header_row, no_col, items)
        today = datetime.now().date().isoformat()
        set_near_label_com(ws, LABEL_RECIPIENT, payload.get("customer") or "")
        set_near_label_com(ws, LABEL_ATTENTION, payload.get("attention") or "")
        set_near_label_com(ws, LABEL_QUOTE_TITLE, payload.get("title") or "")
        set_near_label_com(ws, LABEL_QUOTE_DATE, today)
        set_near_label_com(ws, LABEL_TOTAL, total)
        workbook.Save()
        return True
    except Exception:
        return False
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=True)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def create_quote(payload: dict[str, Any]) -> Path:
    items = payload.get("items") or []
    if not items:
        raise SystemExit("Draft has no items.")

    template = find_template(prefer_vat_split=payload.get("tax_mode", "vat_split") == "vat_split")
    EXPORT_ROOT.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    customer = safe_filename(str(payload.get("customer") or "customer"))
    title = safe_filename(str(payload.get("title") or "quote"))
    output = EXPORT_ROOT / f"{now:%Y%m%d_%H%M%S}_{customer}_{title}.xlsx"
    shutil.copy2(template, output)

    if create_quote_with_excel_com(output, payload, items):
        return output

    wb = load_workbook(output)
    ws = wb.active
    header = find_header(ws)
    if not header:
        wb.close()
        output.unlink(missing_ok=True)
        raise SystemExit("Template has no item table header.")

    header_row, no_col = header
    clear_old_items(ws, header_row, no_col)
    total = write_items(ws, header_row, no_col, items)

    today = now.date().isoformat()
    set_near_label(ws, LABEL_RECIPIENT, payload.get("customer") or "")
    set_near_label(ws, LABEL_ATTENTION, payload.get("attention") or "")
    set_near_label(ws, LABEL_QUOTE_TITLE, payload.get("title") or "")
    set_near_label(ws, LABEL_QUOTE_DATE, today)
    set_near_label(ws, LABEL_TOTAL, total)

    wb.save(output)
    wb.close()
    return output


def archive_quote(output: Path, payload: dict[str, Any]) -> Path:
    customer = safe_filename(str(payload.get("customer") or "customer"))
    archive_dir = RAW_QUOTE_ROOT / customer
    archive_dir.mkdir(parents=True, exist_ok=True)
    archive_path = archive_dir / output.name
    shutil.copy2(output, archive_path)
    return archive_path


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: create_quote_from_draft.py draft.json")

    payload_path = Path(sys.argv[1])
    with payload_path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    output = create_quote(payload)
    archived = archive_quote(output, payload)
    print(json.dumps({"output": str(output), "archived": str(archived)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
